from flask import Flask, jsonify, request
from flask_cors import CORS
from openpyxl import load_workbook
from io import BytesIO

app = Flask(__name__)
CORS(app)

@app.route('/', methods=['GET'])
def home_test():
    return jsonify("conversion up")

@app.route('/excel-to-json', methods=['POST'])
def excel_to_json(): 
    if 'file' not in request.files:
        return jsonify({"error" : "No file part"}), 400
    file = request.files['file']
    if file.filename == '' : 
        return jsonify({"error" : "No file part"}), 400
    if file and file.filename.endswith('.xlsx'): 
        in_memory_file = BytesIO(file.read())
        worbook = load_workbook(in_memory_file)
        json_data = {}
        for sheet in worbook.worksheets : 
            sheet_data = []
            for i in range(1, sheet.max_row):
                row = {}
                for j in range(1, sheet.max_column+1):
                    column_name = sheet.cell(row=1, column=j)
                    row_data = sheet.cell(row=i+1, column=j)
                    row.update({column_name.value: row_data.value})
                sheet_data.append(row)

            json_data[sheet.title] = sheet_data
        return jsonify(json_data)
    return jsonify({'error': 'Invalid file format'}), 400

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001)